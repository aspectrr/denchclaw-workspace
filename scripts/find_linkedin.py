#!/usr/bin/env python3
"""
LinkedIn Profile Finder Agent

Finds LinkedIn profiles for leads from the Elastic Forum Leads spreadsheet.
Can be run standalone or used as a library by a sub-agent.

Usage:
    python3 find_linkedin.py                    # Process next unprocessed lead
    python3 find_linkedin.py --all             # Process all unprocessed leads
    python3 find_linkedin.py --row 3           # Process specific Excel row
    python3 find_linkedin.py --spreadsheet <path>  # Custom spreadsheet path
    python3 find_linkedin.py --status          # Show summary of all leads
"""

import argparse
import os
import sys
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.cell.cell import MergedCell
from dataclasses import dataclass, field
from typing import Optional


DEFAULT_SPREADSHEET = os.path.join(os.path.dirname(__file__), "elastic_forum_leads.xlsx")

COLOR_MAP = {
    "00E2EFDA": "E2EFDA",
    "00FFF2CC": "FFF2CC",
    "00FCE4EC": "FCE4EC",
}


@dataclass
class Lead:
    row: int
    username: str
    name: str
    pain_point: str
    topic_title: str
    forum_link: str
    post_date: str
    last_active: str
    priority: str
    notes: str
    linkedin_value: Optional[str] = None

    @property
    def has_name(self) -> bool:
        return bool(self.name) and self.name.strip() != "" and self.name != self.username

    @property
    def full_name(self) -> Optional[str]:
        if self.has_name:
            return self.name.strip()
        return None

    def is_processed(self) -> bool:
        val = self.linkedin_value
        if val is None:
            return False
        if isinstance(val, str) and val.startswith("http"):
            return True
        if isinstance(val, str) and val.startswith("Not Found"):
            return True
        return False


def read_leads(path: str) -> list[Lead]:
    df = pd.read_excel(path, sheet_name="Elastic Forum Leads")
    leads = []
    for excel_row, (_, row) in enumerate(df.iterrows(), start=2):
        linkedin_raw = row.get("LinkedIn Link", None)
        linkedin_val: Optional[str] = None
        if linkedin_raw is not None and not (isinstance(linkedin_raw, float)):
            linkedin_val = str(linkedin_raw)

        def safe_str(val) -> str:
            return str(val).strip() if val is not None and not (isinstance(val, float)) else ""

        username_val = safe_str(row.get("Username", ""))
        name_val = safe_str(row.get("Name", ""))
        pain_val = safe_str(row.get("Pain Point / Context", ""))
        topic_val = safe_str(row.get("Topic Title", ""))
        forum_val = safe_str(row.get("Forum Link", ""))
        post_val = safe_str(row.get("Post Date", ""))
        active_val = safe_str(row.get("Last Active on Forum", ""))
        priority_val = safe_str(row.get("Outreach Priority", ""))
        notes_val = safe_str(row.get("Notes", ""))

        leads.append(Lead(
            row=excel_row,
            username=username_val,
            name=name_val,
            pain_point=pain_val,
            topic_title=topic_val,
            forum_link=forum_val,
            post_date=post_val,
            last_active=active_val,
            priority=priority_val,
            notes=notes_val,
            linkedin_value=linkedin_val,
        ))
    return leads


def write_result(path: str, row: int, value: str):
    wb = load_workbook(path)
    sheet = wb["Elastic Forum Leads"]
    ref_cell = sheet.cell(row=row, column=1)
    fill_color = ref_cell.fill.start_color.rgb
    fill_hex = COLOR_MAP.get(fill_color, "FFFFFF")
    cell = sheet.cell(row=row, column=10)
    if isinstance(cell, MergedCell):
        print(f"WARNING: Cell J{row} is part of a merged range. Skipping write.")
        return
    cell.value = value
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.border = Border(bottom=Side(style="thin"))
    wb.save(path)


def print_lead(lead: Lead):
    print(f"\n{'='*60}")
    print(f"ROW {lead.row}: {lead.name} (@{lead.username})")
    print(f"Priority: {lead.priority}")
    print(f"Topic: {lead.topic_title}")
    pc = lead.pain_point
    print(f"Context: {pc[:150]}{'...' if len(pc) > 150 else ''}")
    print(f"{'='*60}")


def print_summary(leads: list[Lead]):
    found = sum(1 for l in leads if l.linkedin_value and l.linkedin_value.startswith("http"))
    not_found = sum(1 for l in leads if l.linkedin_value and l.linkedin_value.startswith("Not Found"))
    pending = sum(1 for l in leads if not l.is_processed())
    total = len(leads)
    print(f"\nTotal leads: {total}")
    print(f"  Found:      {found}")
    print(f"  Not Found:  {not_found}")
    print(f"  Pending:    {pending}")
    print(f"\nPending leads:")
    for l in leads:
        if not l.is_processed():
            print(f"  Row {l.row}: {l.name} (@{l.username}) - {l.priority}")


def extract_forum_username_from_link(link: str) -> Optional[str]:
    m = re.search(r'discuss\.elastic\.co/u/([^/?#]+)', link)
    if m:
        return m.group(1)
    return None


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Find LinkedIn profiles for Elastic Forum leads")
    parser.add_argument("--all", action="store_true", help="Process all unprocessed leads")
    parser.add_argument("--row", type=int, help="Process specific Excel row")
    parser.add_argument("--spreadsheet", default=DEFAULT_SPREADSHEET, help="Path to spreadsheet")
    parser.add_argument("--status", action="store_true", help="Show status of all leads")
    args = parser.parse_args()

    if not os.path.exists(args.spreadsheet):
        print(f"ERROR: Spreadsheet not found: {args.spreadsheet}")
        sys.exit(1)

    leads = read_leads(args.spreadsheet)

    if args.status:
        print_summary(leads)
        sys.exit(0)

    if args.row:
        filtered = [l for l in leads if l.row == args.row and not l.is_processed()]
        if not filtered:
            print(f"No unprocessed lead found at row {args.row}")
            sys.exit(1)
        to_process = filtered
    elif args.all:
        to_process = [l for l in leads if not l.is_processed()]
        if not to_process:
            print("No unprocessed leads found.")
            sys.exit(0)
    else:
        unprocessed = [l for l in leads if not l.is_processed()]
        if not unprocessed:
            print("No unprocessed leads found.")
            sys.exit(0)
        to_process = unprocessed[:1]

    print(f"Processing {len(to_process)} lead(s):")
    for lead in to_process:
        print_lead(lead)
        print(f"\nAGENT INSTRUCTIONS:")
        print(f"1. Use the LinkedIn MCP search_people tool or web search to find this person's LinkedIn profile")
        print(f"2. Cross-reference name, role, and technical context to verify the match")
        print(f"3. Update spreadsheet at row {lead.row}, column J with:")
        print(f"   - LinkedIn URL if found (e.g. https://linkedin.com/in/name-here)")
        print(f"   - 'Not Found - <reason>' if not found")
        print(f"   - Run: write_result('{args.spreadsheet}', {lead.row}, '<value>')")
